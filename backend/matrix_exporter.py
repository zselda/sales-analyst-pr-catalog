"""
Product Recommendation Matrix Exporter
=======================================
Extracts the PRODUCT RECOMMENDATION MATRIX TABLE produced by the
strategist agent (Section 3 of the markdown report) into a pandas
DataFrame and saves it as a styled Excel output file — alongside the
other pipeline outputs (Report.html, Report.pdf, network graph.html).

Public API:
    extract_recommendation_matrix(report_md) -> pd.DataFrame | None
    save_product_matrix_excel(result, output_dir, company_name) -> str | None
"""

import os
import re
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("swarm.matrix_exporter")

# ING brand colors (aligned with pdf_generator / html_generator)
ING_NAVY = "000066"
ING_ORANGE = "FF6200"
ING_GRAY = "F5F5F5"

# Heading patterns that mark the matrix section (EN + TR translated report)
_HEADING_PATTERNS = [
    re.compile(r"PRODUCT\s+RECOMMENDATION\s+MATRIX", re.IGNORECASE),
    re.compile(r"ÜRÜN\s+(ÖNERİ|ÖNERİLERİ|TAVSİYE)\s+MATRİS", re.IGNORECASE),
    re.compile(r"ÜRÜN\s+(ÖNERİ|ÖNERİLERİ|TAVSİYE)\s+MATRIS", re.IGNORECASE),
]


def _clean_cell(cell: str) -> str:
    """Strip markdown/HTML artifacts the LLM uses inside table cells."""
    text = re.sub(r"<br\s*/?>", "\n", cell, flags=re.IGNORECASE)
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)          # bold markers
    text = re.sub(r"(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)", r"\1", text)  # italics
    text = text.replace("`", "")
    text = re.sub(r"\s*•\s*", "\n• ", text)               # bullets on own lines
    text = re.sub(r"[ \t]*\n[ \t]*", "\n", text)
    return text.strip().lstrip("\n")


def _split_table_row(line: str) -> list:
    """Split a markdown table row, preserving EMPTY cells (no column shift)."""
    inner = line.strip()
    if inner.startswith("|"):
        inner = inner[1:]
    if inner.endswith("|"):
        inner = inner[:-1]
    return [c.strip() for c in inner.split("|")]


def _is_separator_row(cells: list) -> bool:
    return all(re.fullmatch(r":?-{2,}:?", c) or c == "" for c in cells)


def _parse_table_at(lines: list, start: int) -> pd.DataFrame:
    """Parse a contiguous markdown table starting at lines[start]."""
    header = _split_table_row(lines[start])
    rows = []
    i = start + 1
    while i < len(lines) and "|" in lines[i]:
        cells = _split_table_row(lines[i])
        if not _is_separator_row(cells):
            # Pad/truncate to header width so ragged LLM rows don't break
            cells = (cells + [""] * len(header))[: len(header)]
            cleaned = [_clean_cell(c) for c in cells]
            if any(cleaned):
                rows.append(cleaned)
        i += 1
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows, columns=[_clean_cell(h) for h in header])


def extract_recommendation_matrix(report_md: str) -> pd.DataFrame:
    """
    Find and parse the Product Recommendation Matrix table in a strategist
    report (markdown). Strategy:
      1. Locate a heading matching 'PRODUCT RECOMMENDATION MATRIX' (or its
         Turkish translation) and take the first table within the next lines.
      2. Fallback: take the first table whose header mentions both a
         product column and an evidence column.
    Returns None when no matrix table can be found.
    """
    if not report_md or not isinstance(report_md, str):
        return None
    lines = report_md.split("\n")

    # 1. Heading-anchored search
    for idx, line in enumerate(lines):
        if any(p.search(line) for p in _HEADING_PATTERNS):
            for j in range(idx + 1, min(idx + 12, len(lines))):
                if lines[j].strip().startswith("|"):
                    df = _parse_table_at(lines, j)
                    if not df.empty:
                        return df
            break

    # 2. Header-content fallback (language-agnostic)
    product_kw = ("product", "ürün")
    evidence_kw = ("evidence", "kanıt", "veri")
    for idx, line in enumerate(lines):
        if (line.strip().startswith("|") and idx + 1 < len(lines)
                and "---" in lines[idx + 1]):
            low = line.lower()
            if any(k in low for k in product_kw) and any(k in low for k in evidence_kw):
                df = _parse_table_at(lines, idx)
                if not df.empty:
                    logger.info("[Matrix] Found matrix via header-content fallback")
                    return df
    logger.warning("[Matrix] No Product Recommendation Matrix table found in report")
    return None


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame,
                 company_name: str, tax_id: str, donem: str, language: str):
    ws = wb.create_sheet(title=sheet_name)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if language == "TR":
        title = f"Ürün Öneri Matrisi — {company_name}"
        meta = f"Vergi No: {tax_id}  |  Tarih: {datetime.now():%d %B %Y}  |  Dönem: {donem}"
    else:
        title = f"Product Recommendation Matrix — {company_name}"
        meta = f"Tax ID: {tax_id}  |  Date: {datetime.now():%d %B %Y}  |  Period: {donem}"

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color=ING_NAVY)
    ws["A2"] = meta
    ws["A2"].font = Font(size=10, color="666673")

    header_row = 4
    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=str(col))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ING_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for ri, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for ci, value in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=str(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if (ri - header_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ING_GRAY)

    # Column widths from content (capped so evidence/reasoning stay readable)
    for ci, col in enumerate(df.columns, start=1):
        longest_line = max(
            [len(line) for v in df.iloc[:, ci - 1].astype(str)
             for line in v.split("\n")] + [len(str(col))]
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(60, max(16, longest_line + 4))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws


def save_product_matrix_excel(result: dict, output_dir: str, company_name: str) -> str:
    """
    Extract the strategist's Product Recommendation Matrix from the EN report
    (and the TR translated report when present) and save it as
    {Company}_Product_Matrix.xlsx with one sheet per language.
    Returns the file path, or None when no matrix table was found.
    """
    tax_id = result.get("tax_id", "—")
    donem_info = result.get("donem_info", {})
    donem = donem_info.get("raw", "—") if isinstance(donem_info, dict) else str(donem_info or "—")

    # Expected row count from the deterministic recommendation catalog
    # (built by product_analyst) — used to detect LLM drift in the matrix.
    catalog = (result.get("product_signals") or {}).get("recommendation_catalog") or {}
    expected_rows = catalog.get("total_rows")

    sheets = []  # (sheet_name, df, language)
    df_en = extract_recommendation_matrix(result.get("strategy_report") or "")
    if df_en is not None and not df_en.empty:
        sheets.append(("Matrix_EN", df_en, "EN"))
        if expected_rows is not None and len(df_en) != expected_rows:
            logger.warning(
                f"[Matrix] ⚠️ ROW-COUNT DRIFT: strategist matrix has {len(df_en)} rows "
                f"but the recommendation catalog mandates {expected_rows} — "
                f"the LLM deviated from the MATRIX STABILITY RULE"
            )
    translated = result.get("translated_report") or ""
    if translated:
        df_tr = extract_recommendation_matrix(translated)
        if df_tr is not None and not df_tr.empty:
            sheets.append(("Matrix_TR", df_tr, "TR"))

    if not sheets:
        logger.warning(
            "[Matrix] Product matrix Excel NOT generated — no matrix table "
            "found in the strategist report(s)"
        )
        return None

    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet
    for sheet_name, df, language in sheets:
        _write_sheet(wb, sheet_name, df, company_name, tax_id, donem, language)

    safe_name = re.sub(r"[^\w\s-]", "", company_name).strip().replace(" ", "_")
    filepath = os.path.join(output_dir, f"{safe_name}_Product_Matrix.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    rows = ", ".join(f"{name}: {len(df)} rows" for name, df, _ in sheets)
    logger.info(f"[Matrix] Saved product recommendation matrix: {filepath} ({rows})")
    return filepath
