"""
Tests — Product Recommendation Matrix exporter + PDF-aligned HTML format
=========================================================================
No LLM required.
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from matrix_exporter import extract_recommendation_matrix, save_product_matrix_excel


EN_REPORT = """# Sales Strategy Report

## 3. PRODUCT SIGNALS & CROSS-SELL OPPORTUNITIES

### PRODUCT RECOMMENDATION MATRIX TABLE
| Client Need | Product | Data Evidence | Reasoning |
|---------|---------------|-------------|--------------------|
| Receivables financing | **ING Faktoring** | • Code: 121 • Volume: ₺35,000,000 | Long collection cycle |
| FX hedging | ING Forward/Opsiyon | • Code: 656 • Volume: ₺12M<br>• Code: 646 • Volume: ₺8M | Unhedged exposure |

## 4. NEXT SECTION
Text after.
"""

TR_REPORT = """# Satış Strateji Raporu

### ÜRÜN ÖNERİ MATRİSİ TABLOSU
| Müşteri İhtiyacı | Ürün | Veri Kanıtı | Gerekçe |
|---------|------|-------------|---------|
| Alacak finansmanı | ING Faktoring | • Kod: 121 • Hacim: ₺35M | Uzun tahsilat süresi |
"""


class TestMatrixExtraction:
    def test_extracts_table_after_heading(self):
        df = extract_recommendation_matrix(EN_REPORT)
        assert list(df.columns) == ["Client Need", "Product", "Data Evidence", "Reasoning"]
        assert len(df) == 2

    def test_cell_cleaning(self):
        df = extract_recommendation_matrix(EN_REPORT)
        assert "**" not in df.iloc[0]["Product"]          # bold stripped
        assert "\n" in df.iloc[1]["Data Evidence"]        # <br> → newline
        evidence = df.iloc[1]["Data Evidence"]
        assert evidence.count("•") == 4                   # all bullets preserved
        assert all(line.startswith("•") for line in evidence.split("\n"))  # one per line

    def test_turkish_heading(self):
        df = extract_recommendation_matrix(TR_REPORT)
        assert df is not None and len(df) == 1
        assert df.columns[0] == "Müşteri İhtiyacı"

    def test_header_content_fallback_without_heading(self):
        report = (
            "# Report\n\nSome text.\n\n"
            "| Client Need | Product | Data Evidence | Reasoning |\n"
            "|---|---|---|---|\n"
            "| Need | ING DBS | • Code: 120 | Dealer network |\n"
        )
        df = extract_recommendation_matrix(report)
        assert df is not None and len(df) == 1

    def test_ragged_rows_padded(self):
        report = (
            "### PRODUCT RECOMMENDATION MATRIX TABLE\n"
            "| A | B | C |\n|---|---|---|\n"
            "| 1 | 2 |\n"            # short row → padded
            "| 1 | 2 | 3 | 4 |\n"    # long row → truncated
        )
        df = extract_recommendation_matrix(report)
        assert df.shape == (2, 3)

    def test_no_table_returns_none(self):
        assert extract_recommendation_matrix("# Report\nNo table.") is None
        assert extract_recommendation_matrix("") is None
        assert extract_recommendation_matrix(None) is None


class TestMatrixExcelOutput:
    @pytest.fixture
    def result(self):
        return {
            "tax_id": "1234567890",
            "donem_info": {"raw": "202506"},
            "strategy_report": EN_REPORT,
            "translated_report": TR_REPORT,
        }

    def test_saves_excel_with_both_language_sheets(self, result, tmp_path):
        path = save_product_matrix_excel(result, str(tmp_path), "Test Co")
        assert path is not None and os.path.exists(path)
        assert path.endswith("Test_Co_Product_Matrix.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert wb.sheetnames == ["Matrix_EN", "Matrix_TR"]
        en = wb["Matrix_EN"]
        assert [c.value for c in en[4]] == ["Client Need", "Product", "Data Evidence", "Reasoning"]
        assert "Test Co" in en["A1"].value
        assert "202506" in en["A2"].value

    def test_returns_none_when_no_matrix(self, tmp_path):
        path = save_product_matrix_excel(
            {"strategy_report": "# Report\nNo table."}, str(tmp_path), "X"
        )
        assert path is None
        assert not list(tmp_path.iterdir())  # nothing written


class TestHtmlPdfFormatAlignment:
    """The HTML report must use the same titles/sections as the PDF report."""

    @pytest.fixture
    def result(self):
        return {
            "company_name": "Test Co", "tax_id": "1234567890",
            "donem_info": {"raw": "202506"},
            "strategy_report": EN_REPORT,
            "translated_report": TR_REPORT,
            "wallet_dict": {"deposit": {"ing": 1000, "other": 9000, "other_count": 3}},
            "account_balances": {"600": {"debit": 5.0, "credit": 1000.5, "balance": 995.5}},
            "network_data": {"stats": {"total_receivables": 5e6, "total_payables": 3e6,
                                       "customer_count": 12, "supplier_count": 8,
                                       "bank_count": 4}},
        }

    def test_en_titles_match_pdf(self, result):
        from html_generator import generate_report_html
        html = generate_report_html(result, language="EN")
        assert "GenAI Sales Analysis Report" in html           # PDF page_title
        assert "Competitor Bank Wallet Share" in html          # PDF section 1
        assert ">Summary<" in html                             # PDF section 2
        assert "Revenue Account Balance Distribution" in html  # PDF section 3
        assert "CAO GenAI Sales Analysis Platform" in html     # PDF footer label
        assert "This report was generated by the CAO GenAI" in html  # PDF disclaimer
        assert "₺1,000.50" in html                             # 2-decimal like PDF

    def test_tr_titles_match_pdf(self, result):
        from html_generator import generate_report_html
        html = generate_report_html(result, language="TR")
        assert "GenAI Satış Analiz Raporu" in html
        assert "Rakip Banka Cüzdan Payı" in html
        assert ">Özet<" in html
        assert "Gelir Hesapları Bakiye Dağılımı" in html
        assert "Bu rapor CAO GenAI" in html

    def test_sections_skipped_without_data_and_safe_donem(self):
        from html_generator import generate_report_html
        # No donem_info, no wallet/balances/network — must not crash (PDF parity)
        html = generate_report_html({"company_name": "X", "strategy_report": "# R\ntext"},
                                    language="EN")
        assert "Competitor Bank Wallet Share" not in html
        assert "Revenue Account Balance Distribution" not in html
        assert ">Summary<" not in html
